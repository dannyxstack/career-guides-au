"""Collect recent official salary history for the project's 46 countries.

The collector uses ILOSTAT's official indicator API as a harmonised baseline and
indexes/copies more detailed national files already present under downloads/{cc}.

Run with the project Conda environment:
    conda run -n career-video python scripts/collect_salary_history.py
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import shutil
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date
from pathlib import Path


REPO = Path(__file__).resolve().parents[1]
DOWNLOADS = REPO / "downloads"
OUTPUT_ROOT = DOWNLOADS / "salary"
QUERY_START_YEAR = 2020
TARGET_START_YEAR = 2021
END_YEAR = 2025
API_URL = "https://rplumber.ilo.org/data/indicator/"
USER_AGENT = "career-guides-au salary-history collector/1.0"


@dataclass(frozen=True)
class Country:
    iso2: str
    iso3: str
    name: str


COUNTRIES = [
    Country("ar", "ARG", "Argentina"),
    Country("at", "AUT", "Austria"),
    Country("au", "AUS", "Australia"),
    Country("be", "BEL", "Belgium"),
    Country("br", "BRA", "Brazil"),
    Country("ca", "CAN", "Canada"),
    Country("ch", "CHE", "Switzerland"),
    Country("cl", "CHL", "Chile"),
    Country("cn", "CHN", "China"),
    Country("cz", "CZE", "Czechia"),
    Country("de", "DEU", "Germany"),
    Country("dk", "DNK", "Denmark"),
    Country("ee", "EST", "Estonia"),
    Country("es", "ESP", "Spain"),
    Country("fi", "FIN", "Finland"),
    Country("fr", "FRA", "France"),
    Country("gr", "GRC", "Greece"),
    Country("hr", "HRV", "Croatia"),
    Country("hu", "HUN", "Hungary"),
    Country("id", "IDN", "Indonesia"),
    Country("ie", "IRL", "Ireland"),
    Country("in", "IND", "India"),
    Country("is", "ISL", "Iceland"),
    Country("it", "ITA", "Italy"),
    Country("jp", "JPN", "Japan"),
    Country("kr", "KOR", "South Korea"),
    Country("lt", "LTU", "Lithuania"),
    Country("lu", "LUX", "Luxembourg"),
    Country("lv", "LVA", "Latvia"),
    Country("mx", "MEX", "Mexico"),
    Country("my", "MYS", "Malaysia"),
    Country("nl", "NLD", "Netherlands"),
    Country("no", "NOR", "Norway"),
    Country("nz", "NZL", "New Zealand"),
    Country("pl", "POL", "Poland"),
    Country("pt", "PRT", "Portugal"),
    Country("ro", "ROU", "Romania"),
    Country("se", "SWE", "Sweden"),
    Country("sg", "SGP", "Singapore"),
    Country("si", "SVN", "Slovenia"),
    Country("sk", "SVK", "Slovakia"),
    Country("th", "THA", "Thailand"),
    Country("tr", "TUR", "Turkey"),
    Country("uk", "GBR", "United Kingdom"),
    Country("us", "USA", "United States"),
    Country("vn", "VNM", "Vietnam"),
]


# Ordered from most convenient for an annual salary chart to least convenient.
INDICATORS = {
    "EAR_EMTM_SEX_OCU_NB_A": ("occupation", "median", "monthly", 1),
    "EAR_EMTA_SEX_OCU_NB_A": ("occupation", "average", "monthly", 2),
    "EAR_EHRM_SEX_OCU_NB_A": ("occupation", "median", "hourly", 3),
    "EAR_EHRA_SEX_OCU_NB_A": ("occupation", "average", "hourly", 4),
    "EAR_EMTM_SEX_ECO_NB_A": ("industry", "median", "monthly", 5),
    "EAR_EMTA_SEX_ECO_NB_A": ("industry", "average", "monthly", 6),
    "EAR_EHRM_SEX_ECO_NB_A": ("industry", "median", "hourly", 7),
    "EAR_EHRA_SEX_ECO_NB_A": ("industry", "average", "hourly", 8),
}

RAW_FIELDS = [
    "ref_area",
    "ref_area.label",
    "source",
    "source.label",
    "indicator",
    "indicator.label",
    "sex",
    "sex.label",
    "classif1",
    "classif1.label",
    "time",
    "obs_value",
    "obs_status",
    "obs_status.label",
    "note_classif",
    "note_classif.label",
    "note_indicator",
    "note_indicator.label",
    "note_source",
    "note_source.label",
]

NORMALIZED_FIELDS = [
    "country_code",
    "country_name",
    "granularity",
    "classification_code",
    "classification_label",
    "year",
    "value",
    "measure",
    "period",
    "unit",
    "currency_code",
    "sex_code",
    "sex_label",
    "source_code",
    "source_label",
    "indicator_code",
    "indicator_label",
    "observation_status",
    "notes",
    "priority",
]

NATIONAL_FILE_RE = re.compile(
    r"salary|salar|wage|earning|remuner|income|pay|lons|vin02001|ispv|ksh",
    re.IGNORECASE,
)
NATIONAL_FILE_EXCLUDE_RE = re.compile(
    r"_by_isco|ilostat|readme|source_url|\.part$", re.IGNORECASE
)
CURRENCY_RE = re.compile(r"Currency:\s*([A-Z]{3})\s*-", re.IGNORECASE)
NATIONAL_EXTRA_FILES = {
    "au": {"63060DO001_202505.xlsx"},
    "de": {"62361-0030_de.xlsx"},
    "sg": {"mom_occupations_and_industries_2025.xlsx"},
}
NATIONAL_SOURCE_URLS = {
    ("au", "63060DO001_202505.xlsx"): "https://www.abs.gov.au/statistics/labour/earnings-and-working-conditions/employee-earnings-and-hours-australia/may-2025",
    ("de", "62361-0030_de.xlsx"): "https://genesis.destatis.de/datenbank/online/statistic/62361/table/62361-0030",
    ("nz", "NewZealand_Employment_Earnings_from_main_wage_and_salary_job_by_occupation_sex_age_group_and_ethnic_group_20092025.csv"): "https://explore.data.stats.govt.nz/vis?tm=Earnings%20from%20main%20wage&pg=0&snb=4",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--countries",
        nargs="*",
        help="Optional ISO-2 subset, for example: --countries au us uk",
    )
    parser.add_argument("--workers", type=int, default=6)
    parser.add_argument(
        "--refresh", action="store_true", help="Redownload existing ILOSTAT CSV files"
    )
    parser.add_argument(
        "--no-national-copy",
        action="store_true",
        help="Only index existing national files instead of copying them",
    )
    return parser.parse_args()


def selected_countries(codes: list[str] | None) -> list[Country]:
    if not codes:
        return COUNTRIES
    wanted = {code.lower() for code in codes}
    result = [country for country in COUNTRIES if country.iso2 in wanted]
    missing = wanted - {country.iso2 for country in result}
    if missing:
        raise SystemExit(f"Unknown country code(s): {', '.join(sorted(missing))}")
    return result


def api_url(country: Country, indicator: str) -> str:
    query = urllib.parse.urlencode(
        {
            "id": indicator,
            "ref_area": country.iso3,
            "timefrom": QUERY_START_YEAR,
            "timeto": END_YEAR,
            "format": ".csv",
            "type": "both",
            "best_source": "yes",
        }
    )
    return f"{API_URL}?{query}"


def fetch_bytes(url: str, attempts: int = 4) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    for attempt in range(1, attempts + 1):
        try:
            with urllib.request.urlopen(request, timeout=180) as response:
                return response.read()
        except (urllib.error.URLError, TimeoutError) as exc:
            if attempt == attempts:
                raise RuntimeError(f"Download failed after {attempts} attempts: {url}") from exc
            time.sleep(2**attempt)
    raise AssertionError("unreachable")


def read_csv_bytes(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text, newline="")))


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def collect_one(country: Country, indicator: str, refresh: bool) -> dict[str, object]:
    raw_dir = OUTPUT_ROOT / country.iso2 / "ilostat"
    raw_path = raw_dir / f"{indicator}.csv"
    url = api_url(country, indicator)

    if raw_path.exists() and not refresh:
        rows = list(csv.DictReader(raw_path.open(encoding="utf-8", newline="")))
    else:
        content = fetch_bytes(url)
        rows = read_csv_bytes(content)
        raw_dir.mkdir(parents=True, exist_ok=True)
        # Keep a header-only file too: it is evidence that the official query returned no rows.
        fields = list(rows[0]) if rows else RAW_FIELDS
        write_csv(raw_path, rows, fields)

    return {
        "country": country,
        "indicator": indicator,
        "url": url,
        "path": raw_path,
        "rows": rows,
    }


def note_text(row: dict[str, str]) -> str:
    parts = [
        row.get("note_classif.label", ""),
        row.get("note_indicator.label", ""),
        row.get("note_source.label", ""),
    ]
    return " | ".join(part.strip() for part in parts if part and part.strip())


def currency_code(row: dict[str, str], period: str) -> str:
    if period not in {"monthly", "hourly"}:
        return ""
    match = CURRENCY_RE.search(note_text(row))
    return match.group(1).upper() if match else "LCU"


def normalize_rows(
    country: Country, indicator: str, rows: list[dict[str, str]]
) -> list[dict[str, object]]:
    granularity, measure, period, priority = INDICATORS[indicator]
    result: list[dict[str, object]] = []
    for row in rows:
        year = row.get("time", "")
        value = row.get("obs_value", "")
        if not year.isdigit() or not value:
            continue
        notes = note_text(row)
        result.append(
            {
                "country_code": country.iso2.upper(),
                "country_name": country.name,
                "granularity": granularity,
                "classification_code": row.get("classif1", ""),
                "classification_label": row.get("classif1.label", ""),
                "year": year,
                "value": value,
                "measure": measure,
                "period": period,
                "unit": f"local currency per {period[:-2] if period.endswith('ly') else period}",
                "currency_code": currency_code(row, period),
                "sex_code": row.get("sex", ""),
                "sex_label": row.get("sex.label", ""),
                "source_code": row.get("source", ""),
                "source_label": row.get("source.label", ""),
                "indicator_code": row.get("indicator", ""),
                "indicator_label": row.get("indicator.label", ""),
                "observation_status": row.get("obs_status.label", ""),
                "notes": notes,
                "priority": priority,
            }
        )
    return result


def chart_ready(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Choose the latest five observations per total-sex classification series."""
    total_sex = [row for row in rows if row["sex_code"] == "SEX_T"]
    useful = [
        row
        for row in total_sex
        if not str(row["classification_code"]).endswith(("_TOTAL", "_X"))
    ]
    best_series_priority: dict[tuple[str, str], int] = {}
    for row in useful:
        key = (str(row["granularity"]), str(row["classification_code"]))
        best_series_priority[key] = min(
            int(row["priority"]), best_series_priority.get(key, 999)
        )
    selected = [
        row
        for row in useful
        if int(row["priority"])
        == best_series_priority[(str(row["granularity"]), str(row["classification_code"]))]
    ]
    available_years: dict[tuple[str, str], list[str]] = {}
    for row in selected:
        key = (str(row["granularity"]), str(row["classification_code"]))
        available_years.setdefault(key, []).append(str(row["year"]))
    latest_years = {
        key: set(sorted(set(years))[-5:]) for key, years in available_years.items()
    }
    latest_five = [
        row
        for row in selected
        if str(row["year"])
        in latest_years[(str(row["granularity"]), str(row["classification_code"]))]
    ]
    # A source can change within a series. Retain duplicate-source rows and expose metadata.
    return sorted(
        latest_five,
        key=lambda row: (
            str(row["granularity"]),
            str(row["classification_code"]),
            str(row["year"]),
            str(row["source_code"]),
        ),
    )


def national_candidates(country: Country) -> list[Path]:
    source_dir = DOWNLOADS / country.iso2
    if not source_dir.exists():
        return []
    return sorted(
        path
        for path in source_dir.iterdir()
        if path.is_file()
        and (
            NATIONAL_FILE_RE.search(path.name)
            or path.name in NATIONAL_EXTRA_FILES.get(country.iso2, set())
        )
        and not NATIONAL_FILE_EXCLUDE_RE.search(path.name)
    )


def copy_national_files(country: Country, copy_files: bool) -> list[dict[str, object]]:
    rows = []
    target_dir = OUTPUT_ROOT / country.iso2 / "national_sources"
    for source in national_candidates(country):
        target = target_dir / source.name
        if copy_files:
            target_dir.mkdir(parents=True, exist_ok=True)
            if not target.exists() or target.stat().st_size != source.stat().st_size:
                shutil.copy2(source, target)
        rows.append(
            {
                "file": source.name,
                "original_path": source.relative_to(REPO).as_posix(),
                "copied_path": target.relative_to(REPO).as_posix() if copy_files else "",
                "bytes": source.stat().st_size,
                "source_url": NATIONAL_SOURCE_URLS.get((country.iso2, source.name), ""),
            }
        )
    original_readme = DOWNLOADS / country.iso2 / "README.md"
    if copy_files and original_readme.exists():
        target_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(original_readme, target_dir / "ORIGINAL_README.md")
    if rows:
        write_csv(
            OUTPUT_ROOT / country.iso2 / "national_sources_manifest.csv",
            rows,
            ["file", "original_path", "copied_path", "bytes", "source_url"],
        )
    return rows


def national_row(
    country: Country,
    code: str,
    label: str,
    year: int,
    value: object,
    measure: str,
    period: str,
    currency: str,
    source_code: str,
    source_label: str,
    indicator_label: str,
    notes: str,
) -> dict[str, object]:
    return {
        "country_code": country.iso2.upper(),
        "country_name": country.name,
        "granularity": "occupation",
        "classification_code": code,
        "classification_label": label,
        "year": str(year),
        "value": value,
        "measure": measure,
        "period": period,
        "unit": f"{currency} per {period[:-2] if period.endswith('ly') else period}",
        "currency_code": currency,
        "sex_code": "TOTAL",
        "sex_label": "Total",
        "source_code": source_code,
        "source_label": source_label,
        "indicator_code": source_code,
        "indicator_label": indicator_label,
        "observation_status": "Official national source",
        "notes": notes,
        "priority": 0,
    }


def build_nz_national_chart(country: Country) -> list[dict[str, object]]:
    files = list((DOWNLOADS / "nz").glob("*Earnings*occupation*20092025.csv"))
    if not files:
        return []
    rows = []
    with files[0].open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if (
                row["Year as at June"] not in {"2021", "2022", "2023", "2024", "2025"}
                or row["Measure"] != "Median Weekly Earnings"
                or row["Sex"] != "Total"
                or row["Age group"] != "total"
                or row["Ethnic group"] != "Total"
                or row["Occupation Code"] == "TOTAL"
                or not row["Value"]
            ):
                continue
            rows.append(
                national_row(
                    country,
                    f"ANZSCO_MAJOR_{row['Occupation Code']}",
                    row["Occupation"],
                    int(row["Year as at June"]),
                    row["Value"],
                    "median",
                    "weekly",
                    "NZD",
                    "STATSNZ_HLFS_OCC_EARNINGS",
                    "Stats NZ Household Labour Force Survey",
                    "Median weekly earnings from main wage and salary job by occupation",
                    "Main wage and salary job; June quarter; total sex, age and ethnicity.",
                )
            )
    return rows


def build_de_national_chart(country: Country) -> list[dict[str, object]]:
    source = DOWNLOADS / "de" / "62361-0030_de.xlsx"
    if not source.exists():
        return []
    from openpyxl import load_workbook

    sheet = load_workbook(source, read_only=True, data_only=True).active
    current_year = None
    rows = []
    for values in sheet.iter_rows(min_row=8, values_only=True):
        code, label = values[0], values[1]
        if isinstance(code, str) and re.fullmatch(r"04/20\d{2}", code):
            current_year = int(code[-4:])
            continue
        if (
            current_year is None
            or not isinstance(code, str)
            or not code.startswith("KB10-")
            or not isinstance(values[10], (int, float))
        ):
            continue
        rows.append(
            national_row(
                country,
                code,
                str(label).strip(),
                current_year,
                values[10],
                "median",
                "monthly",
                "EUR",
                "DESTATIS_62361_0030",
                "Destatis GENESIS-Online",
                "Median gross monthly earnings excluding special payments by occupation",
                "April reference month; total sex; German KldB 2010 classification.",
            )
        )
    return rows


def build_au_national_chart(country: Country) -> list[dict[str, object]]:
    source = DOWNLOADS / "au" / "63060DO001_202505.xlsx"
    if not source.exists():
        return []
    from openpyxl import load_workbook

    sheet = load_workbook(source, read_only=True, data_only=True)["Table_5"]
    code_by_label = {
        "Managers": "ANZSCO_MAJOR_1",
        "Professionals": "ANZSCO_MAJOR_2",
        "Technicians and trades workers": "ANZSCO_MAJOR_3",
        "Community and personal service workers": "ANZSCO_MAJOR_4",
        "Clerical and administrative workers": "ANZSCO_MAJOR_5",
        "Sales workers": "ANZSCO_MAJOR_6",
        "Machinery operators and drivers": "ANZSCO_MAJOR_7",
        "Labourers": "ANZSCO_MAJOR_8",
    }
    rows = []
    for values in sheet.iter_rows(min_row=19, max_row=26, values_only=True):
        label, value = values[0], values[5]
        if label not in code_by_label or not isinstance(value, (int, float)):
            continue
        rows.append(
            national_row(
                country,
                code_by_label[label],
                label,
                2025,
                value,
                "average",
                "weekly",
                "AUD",
                "ABS_EEH_6306_2025",
                "Australian Bureau of Statistics",
                "Average weekly total cash earnings by occupation",
                "All employees; all rates of pay; May 2025 reference period.",
            )
        )
    return rows


def build_national_chart(country: Country) -> list[dict[str, object]]:
    builders = {
        "au": build_au_national_chart,
        "de": build_de_national_chart,
        "nz": build_nz_national_chart,
    }
    builder = builders.get(country.iso2)
    return builder(country) if builder else []


def year_summary(rows: list[dict[str, object]]) -> str:
    years = sorted({str(row["year"]) for row in rows})
    return ", ".join(years) if years else "None"


def write_country_readme(
    country: Country,
    raw_results: list[dict[str, object]],
    normalized: list[dict[str, object]],
    chart_rows: list[dict[str, object]],
    national_chart_rows: list[dict[str, object]],
    national_files: list[dict[str, object]],
) -> None:
    occupation = [row for row in chart_rows if row["granularity"] == "occupation"]
    industry = [row for row in chart_rows if row["granularity"] == "industry"]
    occ_codes = {str(row["classification_code"]) for row in occupation}
    ind_codes = {str(row["classification_code"]) for row in industry}
    raw_lines = []
    for result in sorted(raw_results, key=lambda item: str(item["indicator"])):
        count = len(result["rows"])
        raw_lines.append(
            f"- `ilostat/{result['indicator']}.csv`: {count} rows; "
            f"[official query]({result['url']})."
        )
    national_lines = []
    for row in national_files:
        source = f" [Official source]({row['source_url']})." if row["source_url"] else ""
        national_lines.append(
            f"- `national_sources/{row['file']}` ({int(row['bytes']):,} bytes), copied from "
            f"`{row['original_path']}`.{source} See `national_sources/ORIGINAL_README.md` "
            "when present for the original provenance notes."
        )
    if not national_lines:
        national_lines = ["- No matching national occupational salary snapshot was already present."]

    body = f"""# {country.name} salary history ({TARGET_START_YEAR}-{END_YEAR})

Downloaded on {date.today().isoformat()}. This directory is a reusable staging area; nothing here has been loaded into the database.

## Coverage

- ILOSTAT normalized rows ({QUERY_START_YEAR}-{END_YEAR}, including fallback year): {len(normalized):,}
- Chart-ready occupational rows: {len(occupation):,} across {len(occ_codes):,} classification codes
- Chart-ready industry rows: {len(industry):,} across {len(ind_codes):,} classification codes
- National chart-ready occupational rows: {len(national_chart_rows):,}
- Years present in chart-ready data: {year_summary(chart_rows)}
- Years present in normalized national chart data: {year_summary(national_chart_rows)}
- Preferred order: monthly median, monthly average, hourly median, hourly average; occupation before industry

`chart_ready_salary_history.csv` contains each classification's latest five available total-sex observations and removes total/unknown buckets. It uses 2020 only when a newer target year is unavailable. It does not annualize hourly or monthly values, because doing so requires country-specific working-time assumptions. `salary_history_2021_2025.csv` is the strict target window; `salary_history_2020_2025.csv` also retains the fallback year.

## ILOSTAT files

Source: [ILOSTAT bulk/API data](https://ilostat.ilo.org/data/bulk/), International Labour Organization. Values are official reported or harmonized statistics, but survey coverage and classification detail differ by country and year.

{chr(10).join(raw_lines)}

## National files already in this repository

These files can provide finer occupational detail than ILOSTAT, but may cover only one year. They are preserved as source snapshots and are not silently merged with the harmonized series.

{chr(10).join(national_lines)}

Where a national parser is available, `national_chart_ready_salary_history.csv` contains a separate normalized extract. It is intentionally not appended to the ILOSTAT chart file because classifications and survey concepts differ.

## Reuse cautions

- Do not join occupations by label alone. Prefer ISCO/national classification codes and an explicit crosswalk.
- Do not mix mean and median, hourly and monthly, or different survey populations in one line without a methodology note.
- A missing year is missing data, not zero salary.
- Forecast values must be stored separately and labelled as estimates; these files contain historical observations only.
"""
    (OUTPUT_ROOT / country.iso2 / "README.md").write_text(body, encoding="utf-8")


def write_root_summary(coverage: list[dict[str, object]]) -> None:
    write_csv(
        OUTPUT_ROOT / "coverage_summary.csv",
        coverage,
        [
            "country_code",
            "country_name",
            "normalized_rows",
            "chart_rows",
            "occupation_codes",
            "industry_codes",
            "years",
            "national_files",
            "national_chart_rows",
            "national_years",
            "effective_years",
            "best_granularity",
            "effective_best_granularity",
        ],
    )
    complete = sum(
        1 for row in coverage if len(str(row["effective_years"]).split(",")) >= 5
    )
    occupational = sum(
        1 for row in coverage if row["effective_best_granularity"] == "occupation"
    )
    gaps = [
        row for row in coverage if len(str(row["effective_years"]).split(",")) < 5
    ]
    gap_lines = [
        f"- `{row['country_code']}` {row['country_name']}: available years "
        f"{row['effective_years'] or 'none'}; best granularity "
        f"{row['effective_best_granularity']}; {row['national_files']} national source files indexed."
        for row in gaps
    ]
    body = f"""# Salary history source archive

This archive contains official salary/earnings observations for {len(coverage)} project countries, targeting the five completed years {TARGET_START_YEAR}-{END_YEAR}. It was generated on {date.today().isoformat()} by `scripts/collect_salary_history.py` and has not been loaded into the database. Because many 2025 national releases are not yet in ILOSTAT, 2020 is downloaded as a fallback so each series can use its latest five published observations.

## What is included

- `{occupational}` countries have at least one chart-ready occupational series after considering ILOSTAT and parsed national files.
- `{complete}` countries have chart-ready observations in five distinct recent years. This counts years, not identical classification coverage across every year.
- `coverage_summary.csv` is the machine-readable inventory.
- Each country directory contains raw official query results, normalized data, a chart-ready subset, national-source snapshots where available, and a provenance README.

## Known publication gaps

The following countries still have fewer than five distinct recent years after combining the harmonized extract with parsed national files. Do not interpolate these gaps in the historical layer.

{chr(10).join(gap_lines) if gap_lines else '- None.'}

## Directory layout

```text
downloads/salary/
  coverage_summary.csv
  {{country-code}}/
    README.md
    salary_history_2020_2025.csv
    salary_history_2021_2025.csv
    chart_ready_salary_history.csv
    national_chart_ready_salary_history.csv
    national_sources_manifest.csv
    ilostat/*.csv
    national_sources/*
```

## Source and methodology

The cross-country baseline comes from the [ILOSTAT bulk download/API facility](https://ilostat.ilo.org/data/bulk/) and uses eight annual indicators: mean/median monthly/hourly earnings by occupation and by economic activity. API queries request the best available source for {QUERY_START_YEAR}-{END_YEAR} and preserve source, observation-status and survey notes.

The chart-ready selection prefers occupational monthly medians, then occupational monthly means, then hourly equivalents; industry series are retained as a fallback. ILOSTAT occupation records are commonly ISCO major groups rather than individual four-digit jobs. The separate national extracts can be finer, such as Germany's KldB 2010 records. No currency conversion, inflation adjustment, annualization, interpolation or forecast is applied here.

## Recommended next processing step

For a website chart, first select a classification series with stable codes and at least three observed years. Convert nominal values to real local-currency values with an official CPI series, then estimate the future five years in a separate output containing model name, training window, scenario and confidence interval. Never append forecasts to the historical source file without an `observation_type` field.
"""
    (OUTPUT_ROOT / "README.md").write_text(body, encoding="utf-8")


def main() -> None:
    args = parse_args()
    countries = selected_countries(args.countries)
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    jobs = [(country, indicator) for country in countries for indicator in INDICATORS]
    grouped: dict[str, list[dict[str, object]]] = {country.iso2: [] for country in countries}
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        futures = {
            executor.submit(collect_one, country, indicator, args.refresh): (country, indicator)
            for country, indicator in jobs
        }
        for future in as_completed(futures):
            country, indicator = futures[future]
            try:
                result = future.result()
            except Exception as exc:
                print(f"[ERROR] {country.iso2.upper()} {indicator}: {exc}", flush=True)
                continue
            grouped[country.iso2].append(result)
            print(
                f"[{country.iso2.upper()}] {indicator}: {len(result['rows'])} rows",
                flush=True,
            )

    coverage: list[dict[str, object]] = []
    for country in countries:
        results = grouped[country.iso2]
        normalized: list[dict[str, object]] = []
        for result in results:
            normalized.extend(
                normalize_rows(country, str(result["indicator"]), result["rows"])
            )
        normalized.sort(
            key=lambda row: (
                int(row["priority"]),
                str(row["classification_code"]),
                str(row["year"]),
                str(row["sex_code"]),
            )
        )
        chart_rows = chart_ready(normalized)
        country_dir = OUTPUT_ROOT / country.iso2
        write_csv(
            country_dir / "salary_history_2020_2025.csv",
            normalized,
            NORMALIZED_FIELDS,
        )
        target_rows = [
            row for row in normalized if int(str(row["year"])) >= TARGET_START_YEAR
        ]
        write_csv(
            country_dir / "salary_history_2021_2025.csv",
            target_rows,
            NORMALIZED_FIELDS,
        )
        write_csv(
            country_dir / "chart_ready_salary_history.csv",
            chart_rows,
            NORMALIZED_FIELDS,
        )
        national_files = copy_national_files(country, not args.no_national_copy)
        national_chart_rows = build_national_chart(country)
        write_csv(
            country_dir / "national_chart_ready_salary_history.csv",
            national_chart_rows,
            NORMALIZED_FIELDS,
        )
        write_country_readme(
            country, results, normalized, chart_rows, national_chart_rows, national_files
        )

        occ_codes = {
            str(row["classification_code"])
            for row in chart_rows
            if row["granularity"] == "occupation"
        }
        ind_codes = {
            str(row["classification_code"])
            for row in chart_rows
            if row["granularity"] == "industry"
        }
        years = sorted({str(row["year"]) for row in chart_rows})
        national_years = sorted({str(row["year"]) for row in national_chart_rows})
        effective_years = sorted(set(years) | set(national_years))
        coverage.append(
            {
                "country_code": country.iso2.upper(),
                "country_name": country.name,
                "normalized_rows": len(normalized),
                "chart_rows": len(chart_rows),
                "occupation_codes": len(occ_codes),
                "industry_codes": len(ind_codes),
                "years": ",".join(years),
                "national_files": len(national_files),
                "national_chart_rows": len(national_chart_rows),
                "national_years": ",".join(national_years),
                "effective_years": ",".join(effective_years),
                "best_granularity": "occupation" if occ_codes else "industry" if ind_codes else "none",
                "effective_best_granularity": "occupation"
                if occ_codes or national_chart_rows
                else "industry"
                if ind_codes
                else "none",
            }
        )
        print(
            f"[{country.iso2.upper()}] ready: {len(chart_rows)} chart rows, "
            f"{len(occ_codes)} occupation codes, {len(ind_codes)} industry codes",
            flush=True,
        )

    write_root_summary(coverage)
    print(f"Done. Coverage summary: {OUTPUT_ROOT / 'coverage_summary.csv'}")


if __name__ == "__main__":
    main()
