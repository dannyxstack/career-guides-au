# -*- coding: utf-8 -*-
"""按国家把官方统计的「职业薪资中位数 / 平均薪资」写入 occupation_salaries（方案①）。

存储方式：作为 occupation_salaries 的一行，salary_band='median' 或 'mean'，
salary_min=salary_max=金额（年薪，本币），experience 作展示 label，
salary_note 记来源/口径。sort_order：median=-1（排各经验档之前）、mean=98（排最后）。
export（_i18n_fields）只读 experience/min/max/note，故无需改导出管线。

幂等：写入前先删除该国该 measure 的「官方」行（不含估算），再插入。

loader 一次解析同时产出两个 measure（源有哪个给哪个）：
  {occ_key: {'median': int, 'mean': int}}, period
各国数据源按现有 occ_code（各国原生码）crosswalk，用 keyfn 归并。

运行：
  python -m scripts.load_salary_median --country CA --csv <path> [--measure median|mean] [--dry]
  python -m scripts.load_salary_median --country CA --fill    # 用区间中值估算补全缺官方中位数者
"""
import sys, os, csv, argparse
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from db.connection import get_cursor

FULLTIME_HOURS = 2080  # 40h/周 × 52 周，用于时薪→年薪年化
US_TOPCODE = 239200    # BLS 2025 年薪顶格：A_MEDIAN/A_MEAN='#' 表示 >= 此值


def _iv(v):
    """转 int 年薪；空/无效返回 None。"""
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def load_ca(csv_path):
    """Job Bank Wages CSV：国家级行（prov='NAT'）的中位数与平均数。

    Annual_Wage_Flag=1 为年薪，=0 为时薪(×2080)；中位数与平均数同口径。
    """
    out, year = {}, ""
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            if r["prov"] != "NAT":
                continue
            mul = FULLTIME_HOURS if r["Annual_Wage_Flag_Salaire_annuel"].strip() != "1" else 1
            rec = {}
            med = _iv(r["Median_Wage_Salaire_Median"].strip() or None)
            avg = _iv(r["Average_Wage_Salaire_Moyen"].strip() or None)
            if med is not None:
                rec["median"] = med * mul
            if avg is not None:
                rec["mean"] = avg * mul
            if rec:
                out[r["NOC_CNP"].replace("NOC_", "").strip()] = rec
                year = year or r["Reference_Period"].strip()
    return out, year


def _us_val(v):
    if v == "#":
        return US_TOPCODE
    return _iv(v) if v not in ("*", "**", None, "") else None


def load_us(xlsx_path):
    """BLS OES national_M2025_dl.xlsx：O_GROUP='detailed' 的 A_MEDIAN / A_MEAN（年薪）。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    it = wb.active.iter_rows(values_only=True)
    hdr = [str(h) for h in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}
    ic, ig, im, ia = idx["OCC_CODE"], idx["O_GROUP"], idx["A_MEDIAN"], idx["A_MEAN"]
    out = {}
    for r in it:
        if r[ig] != "detailed":
            continue
        rec = {}
        mv, av = _us_val(r[im]), _us_val(r[ia])
        if mv is not None:
            rec["median"] = mv
        if av is not None:
            rec["mean"] = av
        if rec:
            out[str(r[ic]).strip()] = rec
    wb.close()
    return out, "2025"


def load_au(xlsx_path):
    """JSA Occupation profiles Table_4：ANZSCO 4 位周中位收入×52 年化（无平均数）。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Table_4"]
    out, started = {}, False
    for r in ws.iter_rows(values_only=True):
        c0 = "" if r[0] is None else str(r[0]).strip()
        if not started:
            started = c0 == "ANZSCO Code"
            continue
        if not c0.isdigit():
            continue
        w = _iv(r[4])
        if w is not None:
            out[c0] = {"median": w * 52}
    wb.close()
    return out, "May 2025"


def load_uk(xlsx_path):
    """ONS ASHE Table 14.7a Full-Time sheet：SOC 4 位年薪 gross 的 Median / Mean。"""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Full-Time"]
    out, ci, mi, ai = {}, None, None, None
    for r in ws.iter_rows(values_only=True):
        if ci is None:
            row = [("" if c is None else str(c).strip()) for c in r]
            if "Code" in row and "Median" in row:
                ci, mi = row.index("Code"), row.index("Median")
                ai = row.index("Mean") if "Mean" in row else None
            continue
        code = "" if r[ci] is None else str(r[ci]).strip()
        if len(code) != 4 or not code.isdigit():
            continue
        rec = {}
        mv = _iv(r[mi])
        if mv is not None:
            rec["median"] = mv
        if ai is not None:
            av = _iv(r[ai])
            if av is not None:
                rec["mean"] = av
        if rec:
            out[code] = rec
    wb.close()
    return out, "2025"


# INE EAES gran grupo 名称(独特子串) -> CNO-11 一位大类码
_ES_GG = [
    ("Directores y gerentes", "1"), ("científicos e intelectuales", "2"),
    ("profesionales de apoyo", "3"), ("contables, administrativos", "4"),
    ("servicios de restauración", "5"), ("sector agrícola", "6"),
    ("industrias manufactureras", "7"), ("Operadores de instalaciones", "8"),
    ("Ocupaciones elementales", "9"), ("Ocupaciones militares", "0"),
]
ES_TABLE = "36846"  # EAES 四年期：Medias y percentiles por gran grupo CNO-11
_ES_TAG = {"50": "median", "Media": "mean"}  # 序列末段 -> measure


def load_es(_path=None):
    """INE Tempus API 表 36846：CNO 大类的 P50(中位数) 与 Media(均值)，全国年薪 gross。"""
    import requests
    url = f"https://servicios.ine.es/wstempus/js/ES/DATOS_TABLA/{ES_TABLE}?nult=1"
    data = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60).json()
    out, period = {}, ""
    for s in data:
        parts = [p.strip() for p in s.get("Nombre", "").split(".") if p.strip()]
        if len(parts) < 5 or parts[-3] != "Total":
            continue
        measure = _ES_TAG.get(parts[-1])
        if not measure:
            continue
        code = next((c for sub, c in _ES_GG if sub in parts[-2]), None)
        if code is None:
            continue
        rec = out.setdefault(code, {})
        if measure in rec:  # 取该大类首个（顶层聚合块）
            continue
        d = (s.get("Data") or [{}])[-1]
        v = d.get("Valor")
        if v is None or v < 0:
            continue
        rec[measure] = int(round(v))
        period = period or str(d.get("Anyo") or "")
    return out, period


_NZ_MEAS = {"Median Weekly Earnings": "median", "Average Weekly Earnings": "mean"}


def load_nz(csv_path):
    """Stats NZ Employment Earnings：ANZSCO 1 位大类的周中位/周均值×52 年化（Total 口径）。"""
    best = {}  # (code,measure) -> (year, weekly)
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            measure = _NZ_MEAS.get(r["Measure"])
            if (not measure or r["Sex"] != "Total"
                    or r["Age group"] != "total" or r["Ethnic group"] != "Total"):
                continue
            code = r["Occupation Code"].strip()
            if not code.isdigit():
                continue
            v = r["Value"].strip()
            if not v:
                continue
            y = int(r["Year as at June"])
            key = (code, measure)
            if key not in best or y > best[key][0]:
                best[key] = (y, float(v))
    out, years = {}, []
    for (code, measure), (y, w) in best.items():
        out.setdefault(code, {})[measure] = int(round(w * 52))
        years.append(y)
    return out, (str(max(years)) if years else "")


def load_de(xlsx_path):
    """Destatis 62361-0030：KldB 2010 的月薪 gross 中位数(col10)/平均数(col12) Insgesamt，×12 年化。

    仅取文件 3 位聚合行（方案 A）；本库 occ_code 为 4 位 → keyfn 取前 3 位归并。
    值列为纯数或 '.'/'/'（无数据/保密，_iv 返回 None 跳过）；质量标记在相邻 flag 列，不取。
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(values_only=True):
        c0 = "" if r[0] is None else str(r[0]).strip()
        if not c0.startswith("KB10-"):
            continue
        code = c0.replace("KB10-", "").strip()
        if len(code) != 3:
            continue
        rec = {}
        med = _iv(r[10]) if len(r) > 10 else None
        mean = _iv(r[12]) if len(r) > 12 else None
        if med is not None:
            rec["median"] = med * 12
        if mean is not None:
            rec["mean"] = mean * 12
        if rec:
            out[code] = rec
    wb.close()
    return out, "2025"


# AU EEH Table_5 大类名(前缀) -> ANZSCO 一位大类码
_AU_GG = [
    ("Managers", "1"), ("Professionals", "2"), ("Technicians and trades", "3"),
    ("Community and personal service", "4"), ("Clerical and administrative", "5"),
    ("Sales workers", "6"), ("Machinery operators", "7"), ("Labourers", "8"),
]


def load_au_mean(xlsx_path):
    """ABS EEH DO001 Table_5：ANZSCO 1 位大类『全体雇员周均总现金收入』(All rates)×52 年化。

    注意口径：全体雇员(含兼职)、周均总现金，与 median(全职)人群不同。
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Table_5"]
    out, insec = {}, False
    for r in ws.iter_rows(values_only=True):
        c0 = "" if r[0] is None else str(r[0]).strip()
        up = c0.upper()
        if up.startswith("AVERAGE WEEKLY TOTAL CASH EARNINGS"):  # 段头(独立单元格)，非表标题
            insec = True
            continue
        if not insec:
            continue
        if up.startswith("AVERAGE AGE") or c0.startswith("All occupations"):
            break
        code = next((g for name, g in _AU_GG if c0.startswith(name)), None)
        w = _iv(r[5]) if len(r) > 5 else None  # 'All rates of pay' 末列
        if code and w is not None:
            out[code] = {"mean": w * 52}
    wb.close()
    return out, "May 2025"


FR_INSEE_CSV = os.path.join(os.path.dirname(__file__), "..", "downloads", "fr",
                            "DS_DERA_PRIVE_ANNUEL_2024_data.csv")


def load_fr_mean(xlsx_path):
    """FR 均值：INSEE 按 PCS 净月薪均值 → 经 ROME→FAP→PCS 对照表简单平均 → ×12 年化。

    xlsx_path = DARES 对照表(.xls, PCS-2003↔ROME-V3↔FAP-2009)。以 FAP 为枢纽正向填充：
    rome→fap、fap→{pcs}；每个 ROME 取其 FAP 下各 PCS 的 INSEE 均值的简单平均。
    PCS 大小写不一（表小写 / INSEE 大写）→ 统一小写匹配。
    """
    import xlrd
    # 1) INSEE：全维汇总(全国/全活动/全年龄/两性/全工时/全规模)的净月薪均值 by PCS(小写)
    insee = {}
    with open(FR_INSEE_CSV, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f, delimiter=";"):
            if (r["GEO"] == "F" and r["ACTIVITY"] == "_T" and r["AGE"] == "_T"
                    and r["SEX"] == "_T" and r["WKTIME"] == "_T" and r["NUMBER_EMPL"] == "_T"
                    and r["DERA_MEASURE"] == "SALAIRE_NET_EQTP_MENSUEL_MOYENNE"):
                v = _iv(r["OBS_VALUE"])
                if v is not None:
                    insee[r["PCS_ESE"].strip().lower()] = v
    # 2) 对照表：rome->fap，fap->{pcs(小写)}
    wb = xlrd.open_workbook(xlsx_path)
    sh = wb.sheet_by_name("Table")
    fap = None
    rome_fap, fap_pcs = {}, {}
    for i in range(4, sh.nrows):
        f0 = str(sh.cell_value(i, 0)).strip()
        if f0:
            fap = f0
        pcs = str(sh.cell_value(i, 2)).strip().lower()
        rome = str(sh.cell_value(i, 4)).strip()
        if pcs and fap:
            fap_pcs.setdefault(fap, set()).add(pcs)
        if rome and fap:
            rome_fap[rome] = fap
    # 3) rome -> 其 FAP 下各 PCS 的 INSEE 均值简单平均 ×12
    out = {}
    for rome, fp in rome_fap.items():
        means = [insee[p] for p in fap_pcs.get(fp, ()) if p in insee]
        if means:
            out[rome] = {"mean": int(round(sum(means) / len(means) * 12))}
    return out, "2024"


# 国家 -> (loader, note模板(zh母本,含{m}{period}), currency, keyfn)
CONF = {
    "CA": (load_ca, "全国全职年薪{m}（来源：加拿大 Job Bank，{period}普查）", "CAD", lambda c: c),
    "US": (load_us, "全国全职年薪{m}（来源：美国 BLS OES {period}）", "USD", lambda c: c),
    "AU": (load_au, "全职周{m}×52 年化（来源：ABS EEH {period}，ANZSCO 4位）", "AUD", lambda c: (c or "")[:4]),
    "NZ": (load_nz, "周{m}×52 年化（来源：Stats NZ {period}，ANZSCO 1位大类）", "NZD", lambda c: (c or "")[:1]),
    "UK": (load_uk, "全职年薪 gross {m}（来源：ONS ASHE {period}，SOC 4位）", "GBP", lambda c: c),
    "ES": (load_es, "全国年薪{m}（来源：INE EAES {period}，CNO 大类）", "EUR", lambda c: (c or "")[:1]),
    "DE": (load_de, "月薪 gross {m}×12 年化（来源：Destatis Verdiensterhebung {period}，KldB 3位）", "EUR", lambda c: (c or "")[:3]),
}
LABELS = {"median": "薪资中位数", "mean": "平均薪资"}
ZH_M = {"median": "中位数", "mean": "均值"}
SORT = {"median": -1, "mean": 98}

# (国家,measure) mean 来自与 median 不同的源时的覆盖：(loader, note模板, currency, keyfn)
OVERRIDE = {
    ("AU", "mean"): (load_au_mean, "全体雇员周均总现金×52 年化（来源：ABS EEH {period}，ANZSCO 大类）", "AUD", lambda c: (c or "")[:1]),
    ("FR", "mean"): (load_fr_mean, "净月薪 FTE 均值×12 年化（来源：INSEE {period}，ROME→FAP→PCS 简单平均）", "EUR", lambda c: c),
}


def write_country(country, csv_path, dry, measure):
    if (country, measure) in OVERRIDE:
        loader, note_tpl, currency, keyfn = OVERRIDE[(country, measure)]
    else:
        loader, note_tpl, currency, keyfn = CONF[country]
    code_map, period = loader(csv_path)
    vals = {k: rec[measure] for k, rec in code_map.items() if measure in rec}
    if not vals:
        print(f"[{country}] 该源无 {measure} 数据，跳过")
        return
    note = note_tpl.format(m=ZH_M[measure], period=period)
    print(f"[{country}][{measure}] 源含该 measure 职业码 {len(vals):,} | 来源 {period}")

    with get_cursor() as cur:
        cur.execute("SELECT id, occ_code FROM occupations WHERE country_code=%s", (country,))
        occs = cur.fetchall()
    rows = [(o["id"], currency, LABELS[measure], v, v, note, SORT[measure], measure)
            for o in occs
            if (v := vals.get(keyfn((o["occ_code"] or "").strip()))) is not None]
    print(f"[{country}][{measure}] DB 职业 {len(occs)} | 命中写入 {len(rows)} | 缺 {len(occs)-len(rows)}")
    if dry:
        print("[dry] 不写库；样本:", [(r[0], r[3]) for r in rows[:5]])
        return
    with get_cursor() as cur:
        cur.execute(
            "DELETE s FROM occupation_salaries s JOIN occupations o ON o.id=s.occupation_id "
            "WHERE o.country_code=%s AND s.salary_band=%s AND s.salary_note NOT LIKE %s",
            (country, measure, "%" + EST_MARK + "%"))
        deleted = cur.rowcount
        cur.executemany(
            "INSERT INTO occupation_salaries "
            "(occupation_id, currency, experience, salary_min, salary_max, salary_note, sort_order, salary_band) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", rows)
    print(f"[{country}][{measure}] 删旧官方 {deleted} | 新写 {len(rows)}")


EST_MARK = "估算"
EST_LABEL = "薪资中位数"
EST_NOTE = "薪资中位数（估算：基于各经验档区间中值）"


def fill_estimates(country, dry):
    """对该国仍缺 median 行的职业，用各经验档区间中点的统计中位数作估算值补全。

    幂等：只删「估算」median 行（note 含 EST_MARK），不动官方值。
    """
    import statistics
    with get_cursor() as cur:
        cur.execute(
            "SELECT o.id, o.currency FROM occupations o WHERE o.country_code=%s "
            "AND NOT EXISTS (SELECT 1 FROM occupation_salaries m "
            "  WHERE m.occupation_id=o.id AND m.salary_band='median' AND m.salary_note NOT LIKE %s)",
            (country, "%" + EST_MARK + "%"))
        occs = cur.fetchall()
        rows = []
        for o in occs:
            cur.execute(
                "SELECT salary_min, salary_max FROM occupation_salaries "
                "WHERE occupation_id=%s AND (salary_band IS NULL OR salary_band NOT IN ('median','mean')) "
                "AND salary_min IS NOT NULL AND salary_max IS NOT NULL", (o["id"],))
            mids = [(float(r["salary_min"]) + float(r["salary_max"])) / 2 for r in cur.fetchall()]
            if not mids:
                continue
            est = int(round(statistics.median(mids)))
            rows.append((o["id"], o["currency"], EST_LABEL, est, est, EST_NOTE, -1, "median"))
    print(f"[{country}] 缺官方 median 且可估算 {len(rows)}")
    if dry:
        print("[dry] 不写库；样本:", [(r[0], r[3]) for r in rows[:5]])
        return
    with get_cursor() as cur:
        cur.execute(
            "DELETE s FROM occupation_salaries s JOIN occupations o ON o.id=s.occupation_id "
            "WHERE o.country_code=%s AND s.salary_band='median' AND s.salary_note LIKE %s",
            (country, "%" + EST_MARK + "%"))
        deleted = cur.rowcount
        cur.executemany(
            "INSERT INTO occupation_salaries "
            "(occupation_id, currency, experience, salary_min, salary_max, salary_note, sort_order, salary_band) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", rows)
    print(f"[{country}] 删旧估算 {deleted} | 新写估算 {len(rows)}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--country", required=True)
    ap.add_argument("--csv", help="官方数据文件（CSV/XLSX）；--fill / ES 无需")
    ap.add_argument("--measure", default="median", choices=["median", "mean"])
    ap.add_argument("--fill", action="store_true", help="用各经验档区间中值估算补全缺官方中位数者")
    ap.add_argument("--dry", action="store_true")
    a = ap.parse_args()
    if a.fill:
        fill_estimates(a.country, a.dry)
    else:
        if a.country not in CONF and (a.country, a.measure) not in OVERRIDE:
            sys.exit(f"未配置 loader：{a.country}/{a.measure}；估算补全请加 --fill")
        if not a.csv and a.country != "ES":  # ES 经 INE API 取数
            sys.exit("官方模式需 --csv")
        write_country(a.country, a.csv, a.dry, a.measure)
