"""KEIS 중장기 인력수급전망 2023-2033 -> occupation_outlook(+meta) for KR.

数据源（downloads/outlook/KR/，见该目录 README）：
- `중장기 인력수급 전망_2023_2033_전망DB(공급,산업,직업).xlsx`
    · sheet `III-4.` = KSCO-7 직업 세분류(4位) 취업자수 2018/2023/2028/2033（단위 천명）
    · sheet `III-3.` = KSCO-7 직업 대/중분류(2位) 취업자수（兜底用）
- `한국고용직업분류(18)-한국표준직업분류(17) 간 연계표.xlsx`
    = KECO2018 세분류(4位) ↔ KSCO-7 세분류(4位) 官方连接表

编码链路（本地 KR 职业为 KECO 5/6位，实为 KECO2018 세분류4位+细分）：
  本地 occ_code[:4] = KECO2018 세분류
    →(连接表 invert)→ KSCO-7 세분류(4位)
      →(III-4)→ 该职业逐锚点就业曲线（优先，세분류粒度）
      缺则 KSCO 세분류[:2]=중분류 →(III-3)→ 중분류曲线（兜底，粗粒度，同组共用）

单位统一到「人」（源千명 ×1000）。锚点：2018/2023=实测(is_projected=0)，2028/2033=预测(1)。
表结构复用 scripts/load_outlook.py（不改表），按 country+source 先删后插，可重复运行。

用法：
  PYTHONIOENCODING=utf-8 python scripts/load_outlook_kr.py --dry-run
  PYTHONIOENCODING=utf-8 python scripts/load_outlook_kr.py
"""
import os, sys, glob, json, argparse, collections
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from db.connection import get_cursor
from scripts.load_outlook import DDL_SERIES, DDL_META, load_local_occ

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
KRDIR = os.path.join(ROOT, "downloads", "outlook", "KR")
SOURCE = "KEIS Mid-term Manpower Projection"
EDITION = "2023-2033"
YEARS = [2018, 2023, 2028, 2033]
PROJ = {2018: 0, 2023: 0, 2028: 1, 2033: 1}


def _f(x):
    if x is None:
        return None
    s = str(x).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _code(x):
    if x is None:
        return None
    s = str(x).strip().replace(".0", "")
    return s if s.isdigit() else None


def load_crosswalk():
    """KECO2018 세분류(4) -> set(KSCO-7 세분류 4)."""
    f = [p for p in glob.glob(os.path.join(KRDIR, "*연계표*.xlsx")) if "18" in p and "17" in p][0]
    ws = openpyxl.load_workbook(f, read_only=True, data_only=True)[
        openpyxl.load_workbook(f, read_only=True).sheetnames[0]]
    keco2ksco = collections.defaultdict(set)
    for r in ws.iter_rows(min_row=5, values_only=True):
        kc, ks = _code(r[6]), _code(r[8])
        if kc and ks and len(kc) == 4 and len(ks) == 4:
            keco2ksco[kc].add(ks)
    return keco2ksco


def load_forecast():
    """(se4 -> {year:천명}), (jung2 -> {year:천명}) from III-4 / III-3."""
    f = [p for p in glob.glob(os.path.join(KRDIR, "*.xlsx"))
         if "2023" in os.path.basename(p) and "2033" in os.path.basename(p)][0]
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    se4 = {}
    ws = wb["III-4."]
    cj = cs = None
    for r in ws.iter_rows(min_row=15, values_only=True):
        j, s, se = _code(r[2]), _code(r[3]), _code(r[4])
        if j:
            cj, cs = j, None
        if s:
            cs = s
        if se and cj and cs is not None:
            code = cj + cs + se
            vals = {y: _f(r[5 + i]) for i, y in enumerate(YEARS)}
            if len(code) == 4 and any(v is not None for v in vals.values()):
                se4[code] = vals

    jung2 = {}
    ws = wb["III-3."]
    cd = None
    for r in ws.iter_rows(min_row=13, values_only=True):
        dae, jung = _code(r[2]), _code(r[3])
        if dae:
            cd = dae
        if jung and cd:
            code = cd + jung
            vals = {y: _f(r[4 + i]) for i, y in enumerate(YEARS)}
            if len(code) == 2 and any(v is not None for v in vals.values()):
                jung2[code] = vals
    return se4, jung2


def sum_series(dicts):
    out = {}
    for y in YEARS:
        vs = [d[y] for d in dicts if d.get(y) is not None]
        out[y] = sum(vs) if vs else None
    return out


def build(local, keco2ksco, se4, jung2):
    series, meta = [], []
    stats = collections.Counter()
    for occ_code in local.get("KR", {}):
        keco4 = str(occ_code)[:4]
        kscos = keco2ksco.get(keco4)
        if not kscos:
            stats["no_crosswalk"] += 1
            continue
        # 优先세분류
        se_hits = [se4[k] for k in kscos if k in se4]
        if se_hits:
            vals = sum_series(se_hits)
            gran, scode = "세분류", ",".join(sorted(k for k in kscos if k in se4))
            stats["se"] += 1
        else:
            jungs = {k[:2] for k in kscos}
            j_hits = [jung2[j] for j in jungs if j in jung2]
            if not j_hits:
                stats["no_forecast"] += 1
                continue
            vals = sum_series(j_hits)
            gran, scode = "중분류", ",".join(sorted(j for j in jungs if j in jung2))
            stats["jung"] += 1

        got = [(y, vals[y]) for y in YEARS if vals.get(y) is not None]
        if not got:
            stats["empty"] += 1
            continue
        for y, v in got:
            series.append(("KR", str(occ_code), "KECO", y, v * 1000.0, PROJ[y], SOURCE, scode))
        e23, e33 = vals.get(2023), vals.get(2033)
        g = ((e33 - e23) / e23 * 100.0) if (e23 and e33) else None
        note = ("KSCO-7 %s curve via KECO2018 crosswalk; anchors 2018/23/28/33" % gran)
        meta.append(("KR", str(occ_code), "KECO", SOURCE, EDITION, scode,
                     2023, 2033, g, None, note))
    return series, meta, stats


def main(dry):
    local = load_local_occ()
    keco2ksco = load_crosswalk()
    se4, jung2 = load_forecast()
    print("crosswalk KECO2018 세분류=%d ; forecast 세분류=%d 중분류=%d"
          % (len(keco2ksco), len(se4), len(jung2)))
    series, meta, stats = build(local, keco2ksco, se4, jung2)
    tot = len(local.get("KR", {}))
    print("KR local occ=%d ; mapped=%d (세분류 %d + 중분류 %d) ; unmapped=%s"
          % (tot, len(meta), stats["se"], stats["jung"],
             {k: stats[k] for k in ("no_crosswalk", "no_forecast", "empty") if stats[k]}))
    print("series rows=%d" % len(series))
    if dry:
        return
    with get_cursor() as cur:
        cur.execute(DDL_SERIES)
        cur.execute(DDL_META)
        cur.execute("DELETE FROM occupation_outlook WHERE country='KR' AND source=%s", (SOURCE,))
        cur.execute("DELETE FROM occupation_outlook_meta WHERE country='KR' AND source=%s", (SOURCE,))
        cur.executemany(
            "INSERT INTO occupation_outlook (country,occ_code,occ_code_type,year,employment,is_projected,source,source_code)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", series)
        cur.executemany(
            "INSERT INTO occupation_outlook_meta (country,occ_code,occ_code_type,source,source_edition,source_code,base_year,end_year,growth_pct,growth_desc,note)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", meta)
    print("[OK] wrote %d series + %d meta rows for KR" % (len(series), len(meta)))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    main(ap.parse_args().dry_run)
