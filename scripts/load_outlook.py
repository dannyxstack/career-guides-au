"""解析各国 downloads/outlook 原始数据 -> occupation_outlook(逐年序列) + occupation_outlook_meta(头条).

对齐本地职业：读 site/src/data/occupations_v2.json 的 (country, occ_code, occ_code_type) 为准。
各国编码差异：
  AU  本地 ANZSCO 6 位 -> JSA 4 位 Unit Group（取前 4 位映射，多个 6 位共享一条曲线）
  US  SOC（带连字符）精确匹配
  CA  NOC 2021 5 位 精确匹配
  UK  SOC2020 4 位 精确匹配（源职业名前 4 位）
单位统一归一到「人」：AU/US 源为千人(×1000)，CA/UK 已是人。

用法：
  python scripts/load_outlook.py --dry-run            # 只报覆盖率，不写库
  python scripts/load_outlook.py                      # 建表 + 全量入库(AU/US/CA/UK)
  python scripts/load_outlook.py --countries US,UK    # 仅指定国
"""
import os, sys, csv, json, argparse, collections
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from db.connection import get_cursor

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DL = os.path.join(ROOT, "downloads", "outlook")
OCC_JSON = os.path.join(ROOT, "site", "src", "data", "occupations_v2.json")

DDL_SERIES = """
CREATE TABLE IF NOT EXISTS occupation_outlook (
  id BIGINT AUTO_INCREMENT PRIMARY KEY,
  country VARCHAR(4) NOT NULL,
  occ_code VARCHAR(32) NOT NULL,
  occ_code_type VARCHAR(16) NOT NULL,
  year SMALLINT NOT NULL,
  employment DOUBLE NULL,
  is_projected TINYINT NOT NULL DEFAULT 1,
  source VARCHAR(64) NOT NULL,
  source_code VARCHAR(32) NULL,
  UNIQUE KEY uq_series (country, occ_code, year),
  KEY idx_cc (country)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
"""
DDL_META = """
CREATE TABLE IF NOT EXISTS occupation_outlook_meta (
  id BIGINT AUTO_INCREMENT PRIMARY KEY,
  country VARCHAR(4) NOT NULL,
  occ_code VARCHAR(32) NOT NULL,
  occ_code_type VARCHAR(16) NOT NULL,
  source VARCHAR(64) NOT NULL,
  source_edition VARCHAR(64) NULL,
  source_code VARCHAR(32) NULL,
  base_year SMALLINT NULL,
  end_year SMALLINT NULL,
  growth_pct DOUBLE NULL,
  growth_desc VARCHAR(64) NULL,
  note VARCHAR(255) NULL,
  UNIQUE KEY uq_meta (country, occ_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
"""


def load_local_occ():
    """{country: {occ_code: occ_code_type}}"""
    d = json.load(open(OCC_JSON, encoding="utf-8"))
    occ = d if isinstance(d, list) else d.get("occupations", d)
    out = collections.defaultdict(dict)
    for o in occ:
        out[o["country"]][str(o["occ_code"])] = o.get("occ_code_type")
    return out


# ---------- 各国解析：返回 (series_rows, meta_rows) ----------
# series_rows: [(occ_code, year, employment, is_projected, source, source_code)]
# meta_rows:   {occ_code: dict(source, source_edition, source_code, base_year, end_year, growth_pct, growth_desc, note)}

def parse_AU(local):
    src, edi = "JSA Employment Projections", "May 2024-2034"
    path = os.path.join(DL, "AU", "employment_projections_2024-2034.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table_6 Occupation Unit Group"]
    unit = {}  # 4-digit -> (e2024,e2029,e2034,g5,g10)
    for r in ws.iter_rows(min_row=9, values_only=True):
        code = r[2]
        if code is None or not str(code).isdigit() or len(str(code)) != 4 or r[1] != "N":
            continue
        unit[str(code)] = (r[5], r[6], r[7], r[9], r[11])
    series, meta = [], {}
    for occ_code in local["AU"]:
        ug = occ_code[:4]
        u = unit.get(ug)
        if not u:
            continue
        e24, e29, e34, g5, g10 = u
        for y, e, proj in ((2024, e24, 0), (2029, e29, 1), (2034, e34, 1)):
            if e is not None:
                series.append((occ_code, y, e * 1000.0, proj, src, ug))
        meta[occ_code] = dict(source=src, source_edition=edi, source_code=ug,
                              base_year=2024, end_year=2034,
                              growth_pct=(g10 * 100.0 if g10 is not None else None),
                              growth_desc=None,
                              note="5yr growth %.1f%%" % (g5 * 100) if g5 is not None else None)
    return series, meta


def parse_US(local):
    src, edi = "BLS Employment Projections", "2024-2034"
    path = os.path.join(DL, "US", "occupation.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table 1.2"]
    series, meta = [], {}
    seen = local["US"]
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[2] != "Line item":
            continue
        code = str(r[1]).strip() if r[1] is not None else None
        if code not in seen:
            continue
        e24, e34, gpct = r[3], r[4], r[8]
        for y, e, proj in ((2024, e24, 0), (2034, e34, 1)):
            if e is not None:
                series.append((code, y, float(e) * 1000.0, proj, src, code))
        meta[code] = dict(source=src, source_edition=edi, source_code=code,
                          base_year=2024, end_year=2034,
                          growth_pct=(float(gpct) if gpct is not None else None),
                          growth_desc=None, note=None)
    return series, meta


def parse_CA(local):
    src, edi = "COPS", "2024-2033"
    path = os.path.join(DL, "CA", "cops_employment_2024-2033_noc2021.csv")
    series, meta = [], {}
    seen = local["CA"]
    with open(path, encoding="latin-1", newline="") as f:
        rd = csv.reader(f)
        header = next(rd)
        years = [int(y) for y in header[3:]]  # 2023..2033
        for row in rd:
            code = row[0].strip()
            if code not in seen:
                continue
            vals = row[3:]
            got = []
            for y, v in zip(years, vals):
                v = v.strip()
                if v in ("", "N/A"):
                    continue
                e = float(v)
                proj = 0 if y <= 2023 else 1
                series.append((code, y, e, proj, src, code))
                got.append((y, e))
            if got:
                (y0, e0), (y1, e1) = got[0], got[-1]
                g = (e1 - e0) / e0 * 100.0 if e0 else None
                meta[code] = dict(source=src, source_edition=edi, source_code=code,
                                  base_year=y0, end_year=y1, growth_pct=g,
                                  growth_desc=None, note="employment persons, annual")
    return series, meta


def parse_UK(local):
    src, edi = "Skills Imperative 2035", "2023 release"
    path = os.path.join(DL, "UK", "skills_imperative_2035_employment_by_occupation.csv")
    seen = local["UK"]
    by = collections.defaultdict(list)  # code -> [(year, jobs)]
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            name = r["occupation_name"].strip()
            code = name.split()[0]
            if not (code.isdigit() and len(code) == 4):
                continue
            if code not in seen:
                continue
            try:
                by[code].append((int(r["time_period"]), float(r["jobs"])))
            except (ValueError, KeyError):
                continue
    series, meta = [], {}
    for code, pts in by.items():
        pts.sort()
        for y, e in pts:
            proj = 0 if y <= 2023 else 1
            series.append((code, y, e, proj, src, code))
        # 头条：以 2024(首个预测年) 为基准到末年
        d = dict(pts)
        by0 = 2024 if 2024 in d else pts[0][0]
        by1 = pts[-1][0]
        g = (d[by1] - d[by0]) / d[by0] * 100.0 if d.get(by0) else None
        meta[code] = dict(source=src, source_edition=edi, source_code=code,
                          base_year=by0, end_year=by1, growth_pct=g, growth_desc=None,
                          note="<=2023 actual, >=2024 projected (approx); 4-digit SOC derived")
    return series, meta


PARSERS = {"AU": parse_AU, "US": parse_US, "CA": parse_CA, "UK": parse_UK}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--countries", default="AU,US,CA,UK")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    ccs = [c.strip().upper() for c in args.countries.split(",") if c.strip()]

    local = load_local_occ()
    all_series, all_meta = [], []
    for cc in ccs:
        series, meta = PARSERS[cc](local)
        matched = len(meta)
        total = len(local.get(cc, {}))
        rows = len(series)
        print(f"[{cc}] matched {matched}/{total} local occ  ({100*matched/total:.0f}%)  series_rows={rows}")
        for occ_code, y, e, proj, s, sc in series:
            all_series.append((cc, occ_code, local[cc][occ_code], y, e, proj, s, sc))
        for occ_code, m in meta.items():
            all_meta.append((cc, occ_code, local[cc][occ_code], m["source"], m["source_edition"],
                             m["source_code"], m["base_year"], m["end_year"],
                             m["growth_pct"], m["growth_desc"], m["note"]))

    print(f"\nTOTAL series={len(all_series)}  meta={len(all_meta)}")
    if args.dry_run:
        print("(dry-run, nothing written)")
        return

    with get_cursor() as cur:
        cur.execute(DDL_SERIES)
        cur.execute(DDL_META)
        for cc in ccs:
            cur.execute("DELETE FROM occupation_outlook WHERE country=%s", (cc,))
            cur.execute("DELETE FROM occupation_outlook_meta WHERE country=%s", (cc,))
        cur.executemany(
            "INSERT INTO occupation_outlook (country,occ_code,occ_code_type,year,employment,is_projected,source,source_code)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", all_series)
        cur.executemany(
            "INSERT INTO occupation_outlook_meta (country,occ_code,occ_code_type,source,source_edition,source_code,base_year,end_year,growth_pct,growth_desc,note)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", all_meta)
    print(f"[OK] wrote {len(all_series)} series + {len(all_meta)} meta rows for {ccs}")


if __name__ == "__main__":
    main()
